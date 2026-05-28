"""
SQL Column-Level Data Lineage Tracer.

Reads SQL queries from queries-py (variables q1, 92, q3,...) and produces 
one Excel file per query in OUTPUT_DIR, tracing every output column back 
to its source tables) and column(s) through any number of CTE / subquery /
UNION Layers.

Run:
    python sql_lineage.py
"""

from __future__ import annotations

import importlib 
import os
import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any

import sqlglot 
from sqlglot import exp
from openpyxl import Workbook
from gpenpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


#
# CONFIGURATION
#

QUERIES_MODULE = "queries"
DIALECT = "default"
OUTPUT_DIR = "Lineage_output"                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                               
SCHEMA_HINT: dict[str, list[str]] | None = None


#
# MODELS
#

@dataclass
class LineageRecord:
    output_col_position: int
    output_col_name: str
    output_expression_sql: str
    expression_type: str
    is_direct_column: bool
    intermediate_cte_path: str
    lineage_depth: int
    union_branch: int | None
    source_schema: str | None
    source_table: str | None
    source_column: str | None
    source_table_alias: str | None
    transform_expression: str
    notes: str = ""

    def to_excel_row(self) -> list:
        return [
            self.output_col_position,
            self.output_col_name,
            self.output_expression_sql,
            self.expression_type,
            "YES" if self.is_direct_column else "NO",
            self.intermediate_cte_path,
            self.lineage_depth,
            self.union_branch if self.union_branch is not None else "",
            self. source_schema or "",
            self.source_table or "",
            self.source_column or "",
            self.source_table_alias or "",
            self.transform_expression,
            self.notes,
        ]


@dataclass 
class CTEInfo:
    name: str
    definition_order: int
    depends_on_ctes: list[str] = field(default_factory=list)
    base_tables: list[str]= field(default_factory=list)
    output_columns: list[str]= field(default_factory=list)
    sql_preview: str = ""


#
# PARSING
#

def parse_sql(sql: str, dialect: str = "default") -> exp.Expression:
    dialect_arg = None if dialect in ("default", "", None) else dialect
    return sqlglot.parse_one(sql,dialect=dialect_arg)


def extract_ctes(ast: exp.Expression) -> dict[str, exp.Expression]:
    """Return {cte_name_lower: inner_select_or_union_node}.

    Reads the WITH clause of the outermost statement only - nested CTES in 
    subqueries are intentionally excluded to keep scope semantics correct.
    """
    cte_map: dict[str, exp.Expression] = {}
    with_clause = ast.args.get("with") or ast.args.get("with_")
    if with_clause is None:
        return cte_map
    for cte in with_clause.expressions:
        name = cte.alias_or_name
        if name:
            cte_map[name.lower()] = cte.this
    return cte_map


def get_top_level_select(ast: exp.Expression) -> exp.Expression:
    """Return the SELECT or UNION that produces the final output."""
    if isinstance(ast, (exp.Select, exp.Union)):
        return ast
    if isinstance(ast, exp.Create) and ast. expression is not None: 
        if isinstance(ast. expression, (exp.Select, exp.Union)) :
            return ast.expression
    inner = ast.this if hasattr(ast, "this") else None
    if isinstance(inner, (exp.Select, exp. Union)):
        return inner
    found = ast.find (exp.Select)
    if found is None:
        raise ValueError("No SELECT found in SQL")
    return found


#
# SCOPE
#

def _table_full_name(t: exp.Table) -> str:
    parts = [p for p in (t.args.get("catalog"), t.args.get("db"), t.this) if p]
    return ".".join(p.name if hasattr(p, "name") else str(p) for p in parts)


def _split_schema_table(full_name: str) -> tuple[str | None, str]:
    parts = full_name.split(".")
    if len (parts) == 1:
        return None, parts[0]
    return ".".join(parts[:-11]), parts[-1]


def build_source_map(
    select_node: exp.Select, 
    cte_map: dict[str, exp.Expression], 
) -> dict[str, tuple[str, Any]]:
    """Map every alias / table reference in FROM + JOINs to its source.

    Returns: Callas_lower: ("table"|"cte" |"subquery", value)}
        - value-is the full table name for "table"
        - value is the CTE name (lowercase) for "cte"
        - value isthe inner Select/Union node for "subquery"
    """
    sources: dict[str, tuple[str, Any]] = {}
    
    from_clause = select_node.args.get("from") or select_node.args.get("from_")
    if from_clause is not None:
        _add_source(from_clause.this, sources, cte_map)
    
    for join in select_node.args.get("joins") or []:
        _add_source(join.this, sources, cte_map)
    
    return sources


def _add_source(
    node: exp.Expression,
    sources: dict[str, tuple[str, Any]],
    cte_map: dict[str, exp.Expression],
) -> None:
    if isinstance(node, exp. Table):
        full = _table_full_name(node)
        alias = (node.alias or node.name).lower()
        if node.name.lower() in cte_map:
            sources[alias] = ("cte", node.name.lower())
        else:
            sources[alias] = ("table", full)
    elif isinstance(node, exp.Subquery) :
        alias = (node.alias or "").lower() or f"_subq_{id(node)}"
        sources[alias] = ("subquery", node.this)
    elif isinstance(node, exp.Alias) :
        _add_source(node.this, sources, cte_map)


#
# EXPRESSION CLASSIFICATION
#

def classify_expression(expr: exp.Expression) -> str:
    if isinstance(expr, exp.Alias) :
        expr = expr.this
    if isinstance(expr, exp.Window):
        return "WINDOW"
    if isinstance(expr, exp.AggFunc) :
        return "AGGREGATE"
    if isinstance(expr, exp.Case) :
        return "CASE"
    if isinstance(expr, exp.Subquery) :
        return "SUBQUERY"
    if isinstance(expr, exp.Star):
        return "STAR"
    if isinstance(expr, exp.Literal) or isinstance(expr, exp.Boolean) or isinstance(expr, exp.Null):
        return "LITERAL"
    if isinstance(expr, exp.Column) :
        return "COLUMN_REF"
    if isinstance(expr, (exp.Add, exp.Sub, exp.Mul, exp.Div, exp.Mod)) :
        return "ARITHMETIC"
    if isinstance(expr, exp.Func):
        return "FUNCTION"
    return "EXPRESSION"


def extract_column_refs(expr: exp.Expression) -> list[exp.Column]:
    """All exp.Column refs inside expr (recursive)."""
    return list (expr. find_all(exp.Column))


def get_output_col_name(expr: exp.Expression, position: int) -> str:
    if isinstance(expr, exp.Alias) :
        return expr.alias
    if isinstance(expr, exp.Column) :
        return expr.name
    if isinstance(expr, exp.Star):
        return "*"
    return f"col_{position}"


def _build_transform(output_expr: exp.Expression, col_ref: exp.Column) -> str:
    full_sql = output_expr.sql()
    col_sql= col_ref.sql()
    if full_sql == col_sql:
        return col_sql
    if len(full_sql) > 100:
        return full_sql[:97] + "..."
    return full_sql


def _column_role_note(output_expr: exp.Expression, col_ref: exp.Column) -> str:
    """Best-effort role annotation: COALESCE arg, WINDOW PARTITION col, etc."."""
    node = col_ref
    crossed_keys: list[str] = []
    while node is not None and node is not output_expr:
        if node.arg_key:
            crossed_keys.append(node.arg_key)
        parent = node.parent
        if parent is None:
            break
        if isinstance(parent, exp.Window):
            if "partition_by" in crossed_keys: 
                return "WINDOW PARTITION col"
            if "order" in crossed_keys:
                return "WINDOW ORDER col"
            return "WINDOW input col"
        if isinstance(parent, exp. Case):
            return "CASE branch col"
        if isinstance(parent, exp.Coalesce):
            return "COALESCE arg"
        node = parent
    return ""


#
# UNION
#

def flatten_union(node: exp.Expression) -> list[exp.Select]:
    """Recursively flatten a Union tree into a left-to-right list of Selects."""
    if isinstance(node, exp.Union):
        return flatten_union(node.left) + flatten_union(node.right)
    return [node]


#
# CORE RESOLUTION
#

def _expand_star_for_source(
    source_key: str, 
    source_type: str,
    source_value: Any, 
    cte_map: dict[str, exp.Expression], 
    schema_hint: dict[str, list[str]] | None,
) -> list[tuple[str | None, str, str]]:
    """Expand a `t.*` or `*` against a single source.

    Returns: [(source_table_full_name_or_cte_name, col_name, alias_used)]
    For a CTE source, recurses into the CTE's SELECT list to enumerate cols.
    For a base table, uses schema_hint if provided, else returns sentinel '*'.
    """
    if source_type == "cte":
        cte_node = cte_map.get(source_value)
        if cte_node is None:
            return [(source_value, "*", source_key)]
        cols = _select_output_names(cte_node)
        return [(source_value, c, source_key) for c in cols]
    if source_type == "subquery":
        cols = _select_output_names(source_value)
        return [(f"subq_{source_key}", c, source_key) for c in cols]
    # base table
    if schema_hint:
        _, table_name = _split_schema_table(source_value)
        if table_name in schema_hint:
            return [(source_value, c, source_key) for c in schema_hint[table_name]]
        if source_value in schema_hint:
            return [(source_value, c, source_key) for c in schema_hint[source_value]]
        return [(source_value, "*", source_key)]


def _select_output_names (node: exp.Expression) -> list[str]:
    """Return the named outputs of a SELECT or UNION node (best-effort)."""
    if isinstance(node, exp.Union):
        branches = flatten_union(node)
        return _select_output_names (branches[0]) if branches else []
    if not isinstance(node, exp.Select):
        return []
    out: list[str] = []
    for i, e in enumerate(node.expressions, 1): 
        if isinstance(e, exp.Star):
            out.append("*")
        else:
            out.append(get_output_col_name(e, 1))
    return out


def _find_unqualified_source(
    col_name: str,
    source_map: dict[str, tuple[str, Any]],
    cte_map: dict[str, exp.Expression], 
    schema_hint: dict[str, list[str]] | None, 
 ) -> tuple[str | None, list[str]]:
    """Try to figure out which source contains an unqualified column.

    Returns (chosen_alias, candidate_aliases). If only one candidate is found, 
    chosen_alias is that source. If multiple, chosen_alias is None and 
    candidates lists them all (for an AMBIGUOUS note).
    """
    candidates: list[str] = []
    for alias, (stype, sval) in source_map.items():
        if stype == "cte":
            cte_node = cte_map.get(sval)
            if cte_node is None:
                continue
            names = [n.lower() for n in _select_output_names(cte_node)]
            if col_name.ower() in names or "*" in names:
                candidates.append(alias)
        elif stype == "subquery":
            names = [n.lower() for n in _select_output_names (sval)]
            if col_name.lower() in names or "*" in names:
                candidates.append(alias)
        else:
            if schema_hint:
                _, table_name = _split_schema_table(sval)
                cols = schema_hint.get(table_name) or schema_hint.get(sval) or []
                if col_name.lower() in [c.lower() for c in cols]:
                    candidates.append(alias)
                    continue
            candidates.append(alias)

    if len(candidates) == 1:
        return candidates[0], candidates
    if len(source_map) == 1:
        only = next(iter(source_map))
        return only, [only]
    return None, candidates


def resolve_column_sources(
    output_expr: exp.Expression, 
    cte_map: dict[str, exp.Expression], 
    source_map: dict[str, tuple[str, Any]],
    depth: int, 
    cte_path: list[str], 
    union_branch: int | None, 
    output_col_name: str, 
    output_col_position: int,
    schema_hint: dict[str, list[str]] | None,
)-> list[LineageRecord]:
    """Resolve every Column ref inside output_expr into LineageRecord(s)."""
    actual_expr = output_expr.this if isinstance(output_expr, exp.Alias) else output_expr
    expr_type = classify_expression(actual_expr)
    expr_sql= actual_expr.sql()
    is_direct = isinstance(actual_expr, exp.Column)

    col_refs = extract_column_refs(actual_expr)

    if not col_refs:
        return [LineageRecord(
            output_col_position=output_col_position,
            output_col_name=output_col_name,
            output_expression_sql=expr_sql,
            expression_type="LITERAL" if expr_type == "LITERAL" else expr_type,
            is_direct_column=False,
            intermediate_cte_path=" -> ".join(cte_path + ["FINAL"]),
            lineage_depth=depth,
            union_branch=union_branch,
            source_schema=None,
            source_table="LITERAL",
            source_column=None,
            source_table_alias=None,
            transform_expression=expr_sql,
            notes="No column references",            
        )]

    records: list[LineageRecord] = []
    for col_ref in col_refs:
        records.extend(_resolve_single_column_ref(
            col_ref=col_ref,
            output_expr=actual_expr,
            cte_map=cte_map,
            source_map=source_map,
            depth=depth,
            cte_path=cte_path,
            union_branch=union_branch,
            output_col_name=output_col_name,
            output_col_position=output_col_position,
            output_expression_sql=expr_sql,
            expression_type=expr_type,
            is_direct=is_direct,
            schema_hint=schema_hint,
        ))
    return records


def _resolve_single_column_ref(
    col_ref: exp.Column, 
    output_expr: exp.Expression, 
    cte_map: dict[str, exp.Expression], 
    source_map: dict[str, tuple[str, Any]],
    depth: int, 
    cte_path: list[str],
    union_branch: int | None, 
    output_col_name: str,
    output_col_position: int, 
    output_expression_sql: str,
    expression_type: str,
    is_direct: bool,
    schema_hint: dict[str, list[str]] | None,
) -> list[LineageRecord]:
    table_qualifier = col_ref.table.lower() if col_ref.table else None
    col_name = col_ref.name
    note =_column_role_note(output_expr, col_ref)

    if table_qualifier and table_qualifier in source_map:
        chosen_alias = table_qualifier
    else:
        chosen_alias, candidates =_find_unqualified_source(
            col_name, source_map, cte_map, schema_hint
        )
        if chosen_alias is None:
            ambig_note = (
                f"AMBIGUOUS: matches {len(candidates)} sources ({', '.join(candidates)})" 
                if candidates else f"Cannot resolve {col_ref.sql()}"
            )
            return [LineageRecord(
                output_col_position=output_col_position,
                output_col_name=output_col_name,
                output_expression_sql=output_expression_sql,
                expression_type=expression_type,
                is_direct_column=is_direct,
                intermediate_cte_path=" -> ".join(cte_path + ["FINAL"]),
                lineage_depth=depth,
                union_branch=union_branch,
                source_schema=None, 
                source_table="UNRESOLVED",
                source_coLumn=col_name,
                source_table_alias=table_qualifier,
                transform_expression=_build_transform(output_expr, col_ref),
                notes=(note + "; " if note else "") + ambig_note,
            )]
    
    source_type, source_value = source_map[chosen_alias]

    if source_type == "table":
        schema, table = _split_schema_table(source_value)
        return [LineageRecord(
            output_col_position=output_col_position,
            output_col_name=output_col_name,
            output_expression_sql=output_expression_sql,
            expression_type=expression_type,
            is_direct_column=is_direct,
            intermediate_cte_path=" -> ".join(cte_path + ["FINAL"]),
            lineage_depth=depth,
            union_branch=union_branch,
            source_schema=schema,
            source_table=table,
            source_coLumn=col_name,
            source_table_alias=chosen_alias,
            transform_expression=_build_transform(output_expr, col_ref),
            notes=note,
        )]

    if source_type == "cte":
        cte_node = cte_map[source_value]
        sub_records = resolve_column_in_subquery(
            col_name=col_name,
            subq_node=cte_node,
            cte_map=cte_map,
            depth=depth + 1,
            cte_path=cte_path + [source_value],
            union_branch=union_branch,
            output_col_name=output_col_name,
            output_col_position=output_col_position,
            output_expression_sql=output_expression_sql,
            expression_type=expression_type,
            is_direct=is_direct,
            schema_hint=schema_hint,
        )

        for r in sub_records:
            if note and note not in r.notes:
                r.notes = (r.notes + "; " + note) if r.notes else note
        return sub_records

    if source_type == "subquery":
        return resolve_column_in_subquery(
            col_name=col_name,
            subq_node=source_value,
            cte_map=cte_map,
            depth=depth + 1,
            cte_path=cte_path + [f"subq_{chosen_alias}"],
            union_branch=union_branch,
            output_col_name=output_col_name,
            output_col_position=output_col_position,
            output_expression_sql=output_expression_sql,
            expression_type=expression_type,
            is_direct=is_direct,
            schema_hint=schema_hint,
        )
    return []


def resolve_column_in_subquery(
    col_name: str,
    subq_node: exp.Expression,
    cte_map: dict[str, exp. Expression],
    depth: int, 
    cte_path: list[str],
    union_branch: int | None,
    output_col_name: str,
    output_col_position: int,
    output_expression_sql: str,
    expression_type: str,
    is_direct: bool,
    schema_hint: dict[str, list[str]] | None,
) -> list[LineageRecord]:
    """Find the expression in subq_node that produces colname, then recurse."""
    if isinstance(subq_node, exp.Union):
        results: list[LineageRecord] = []
        for branch in flatten_union(subq_node):
            results.extend(resolve_column_in_subquery(
                col_name, branch, cte_map, depth, cte_path, 
                union_branch, output_col_name, output_col_position, 
                output_expression_sql, expression_type, is_direct, schema_hint,
            ))
        return results

    if not isinstance(subq_node, exp.Select) :
        return [_unresolved(
            output_col_position, output_col_name, output_expression_sql, 
            expression_type, is_direct, cte_path, depth, union_branch, 
            col_name, f"Unsupported node type {type(subq_node).__name__}",
        )]
    
    sub_source_map = build_source_map(subq_node, cte_map)

    for e in subq_node.expressions:
        inner = e.this if isinstance(e, exp.Alias) else e
        inner_name = (
            e.alias if isinstance(e, exp.Alias)
            else (e.name if isinstance(e, exp.Column) else None)
        )
        if isinstance(e, exp.Star):
            results: list[LineageRecord] = []
            for alias_key, (stype, sval) in sub_source_map.items():
                expanded = _expand_star_for_source(
                    alias_key, stype, sval, cte_map, schema_hint
                )
                for src_full, sc, src_alias in expanded:
                    if sc.lower() == col_name.lower() or sc == "*":
                        if stype == "table":
                            schema, table = _split_schema_table(src_full)
                            results.append(LineageRecord(
                                output_col_position=output_col_position, 
                                output_col_name=output_col_name,
                                output_expression_sql=output_expression_sql,
                                expression_type=expression_type,
                                is_direct_column=is_direct,
                                intermediate_cte_path=" -> ".join(cte_path + ["FINAL"]),
                                lineage_depth=depth,
                                union_branch=union_branch,
                                source_schema=schema,
                                source_table=table,
                                source_column=col_name if sc == "*" else sc,
                                source_table_alias=src_alias,
                                transform_expression=col_name,
                                notes="passthrough via SELECT *",
                            ))
                        elif stype == "cte":
                            results.extend(resolve_column_in_subquery(
                                col_name=col_name if sc == "*" else sc,
                                subq_node=cte_map[sval],
                                cte_map=cte_map,
                                depth=depth + 1, 
                                cte_path=cte_path + [sval],
                                union_branch=union_branch,
                                output_col_name=output_col_name,
                                output_col_position=output_col_position,
                                output_expression_sql=output_expression_sql,
                                expression_type=expression_type,
                                is_direct=is_direct,
                                schema_hint=schema_hint,
                            ))
                        elif stype == "subquery":
                            results.extend(resolve_column_in_subquery(
                                col_name=col_name if sc == "*" else sc,
                                subq_node=sval,
                                cte_map=cte_map,
                                depth=depth + 1,
                                cte_path=cte_path + [f"subq_{alias_key}"],
                                union_branch=union_branch,
                                output_col_name=output_col_name,
                                output_col_position=output_col_position,
                                output_expression_sql=output_expression_sql,
                                expression_type=expression_type,
                                is_direct=is_direct,
                                schema_hint=schema_hint,
                            ))
            if results:
                return results
            continue

        if inner_name and inner_name.lower() == col_name.lower():
            return resolve_column_sources(
                output_expr=inner,
                cte_map=cte_map,
                source_map=sub_source_map,
                depth=depth,
                cte_path=cte_path,
                union_branch=union_branch,
                output_col_name=output_col_name,
                output_col_position=output_col_position,
                schema_hint=schema_hint,
            )
    
    return [_unresolved(
        output_col_position, output_col_name, output_expression_sql, 
        expression_type, is_direct, cte_path, depth, union_branch, 
        col_name, f"col '{col_name}' not found in subquery SELECT list",
    )]


def _unresolved(
    pos, name, expr_sql, expr_type, is_direct, 
    cte_path, depth, union_branch, col_name, note,
) -> LineageRecord:
    return LineageRecord(
        output_col_position=pos,
        output_col_name=name, 
        output_expression_sql=expr_sql,
        expression_type=expr_type,
        is_direct_column=is_direct,
        intermediate_cte_path=" -> ".join(cte_path + ["FINAL"]),
        lineage_depth=depth,
        union_branch=union_branch,
        source_schema=None,
        source_table="UNRESOLVED",
        source_column=col_name,
        source_table_alias=None,
        transform_expression="",
        notes=note,
    )


#
# TOP-LEVEL PROCESSING
#

def process_final_select(
    select_node: exp.Select,
    cte_map: dict[str, exp.Expression], 
    union_branch: int | None,
    schema_hint: dict[str, list[str]] | None,
) -> list[LineageRecord]:
    source_map = build_source_map(select_node, cte_map)
    records: list[LineageRecord] = []

    for position, expr in enumerate(select_node.expressions, 1):
        if isinstance(expr, exp.Star):
            for alias_key, (stype, sval) in source_map.items() :
                expanded = _expand_star_for_source(
                    alias_key, stype, sval, cte_map, schema_hint
                )
                for src_full, col_name, src_alias in expanded:
                    if stype == "table":
                        schema, table = _split_schema_table(src_full)
                        records.append(LineageRecord(
                            output_col_position=position,
                            output_col_name=col_name,
                            output_expression_sql="*",
                            expression_type="STAR",
                            is_direct_column=True,
                            intermediate_cte_path="FINAL",
                            lineage_depth=0,
                            union_branch=union_branch,
                            source_schema=schema,
                            source_table=table,
                            source_column=col_name,
                            source_table_alias=src_alias,
                            transform_expression="*",
                            notes=("schema not provided - cannot expand"
                                    if col_name == "*" else ""),
                        ))
                    else:
                        sub_recs = resolve_column_in_subquery(
                            col_name=col_name,
                            subq_node=(cte_map[sval] if stype == "cte" else sval),
                            cte_map=cte_map,
                            depth=1,
                            cte_path=[sval if stype == "cte" else f"subq_{alias_key}"],
                            union_branch=union_branch,
                            output_col_name=col_name,
                            output_col_position=position, 
                            output_expression_sql="*",
                            expression_type="STAR",
                            is_direct=True,
                            schema_hint=schema_hint,
                        )
                        records.extend(sub_recs)
            continue

        output_name = get_output_col_name(expr, position)
        records.extend(resolve_column_sources(
            output_expr=expr,
            cte_map=cte_map,
            source_map=source_map,
            depth=0,
            cte_path=[],
            union_branch=union_branch,
            output_col_name=output_name,
            output_col_position=position,
            schema_hint=schema_hint,
        ))
    return records


def analyze_cte_dependencies(
    cte_map: dict[str, exp.Expression],
) -> list[CTEInfo]:
    infos: list[CTEInfo] = []
    for order, (name, node) in enumerate(cte_map.items(), 1):
        depends: list[str] = []
        base: list[str] = []
        for tbl in node.find_all(exp.Table):
            tname = tbl.name.lower()
            if tname in cte_map and tname != name:
                if tname not in depends: 
                    depends.append(tname)
            else:
                full = _table_full_name(tbl)
                if full not in base:
                    base.append(full)
        outputs = _select_output_names(node)
        sql_preview = node.sql()[:300]
        infos.append(CTEInfo(
            name=name,
            definition_order=order,
            depends_on_ctes=depends,
            base_tables=base,
            output_columns=outputs,
            sql_preview=sql_preview,
        ))
    return infos


#
# EXCEL WRITER
#

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
UNRESOLVED_FILL = PatternFill("solid", fgColor="FFD580")
NON_DIRECT_FILL = PatternFill("solid", fgColor="FFF9C4")

LINEAGE_HEADERS = [
    "Output Col #", "Output Column Name", "Output Expression",
    "Expression Type", "Is Direct Passthrough", "CTE / Subquery Path",
    "Lineage Depth", "Union Branch", "Source Schema", "Source Table",
    "Source Column", "Source Table Alias", "Transform Expression", "Notes",
]
CTE_HEADERS = [
    "CTE Name", "Definition Order", "Depends On CTEs",
    "Base Tables Used", "Output Columns", "SQL Preview",
]
META_HEADERS = ["Metr1c", "Value"]


def _autosize(ws, min_w=8, max_w=60):
    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        max_len = 0
        for c in col_cells:
            v = c.value
            if v is None:
                continue
            s = str(v)
            if "\n" in s:
                s = max(s.split("\n"), key=len)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[letter].width = max(min_w, min(max_len + 2, max_w))


def write_excel(
    records: list[LineageRecord], 
    cte_infos: list[CTEInfo],
    output_path: str,
    sql: str, 
    dialect: str,
) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Lineage"

    for col, h in enumerate(LINEAGE_HEADERS, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical="center")

    for row_idx, rec in enumerate(records, start=2):
        row_values = rec.to_excel_row()
        is_unresolved = rec.source_table == "UNRESOLVED"
        is_alt = (row_idx % 2) == 0
        fill = None
        if is_unresolved:
            fill = UNRESOLVED_FILL
        elif not rec.is_direct_column:
            fill = NON_DIRECT_FILL
        elif is_alt:
            fill = ALT_FILL
        for col_idx, val in enumerate(row_values, 1):
            c = ws1.cell(row=row_idx, column=col_idx, value=val)
            if fill is not None:
                c.fill = fill
            c.alignment = Alignment(vertical="top", wrap_text=False)

    ws1. freeze_panes = "A2"
    if records:
        last_col = get_column_letter(len(LINEAGE_HEADERS))
        ws1.auto_filter.ref = f"A1:{last_col}{len(records) + 1}"
    _autosize(ws1)

    ws2 = wb. create_sheet("CTE_Map")
    for col, h in enumerate(CTE_HEADERS, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    for row_idx, info in enumerate(cte_infos, start=2): 
        ws2.cell(row=row_idx, column=1, value=info.name)
        ws2.cell(row=row_idx, column=2, value=info.definition_order)
        ws2.cell(row=row_idx, column=3, value=", ".join(info.depends_on_ctes))
        ws2.cell(row=row_idx, column=4, value=", ".join(info.base_tables))
        ws2.cell(row=row_idx, column=5, value=", ".join(info.output_columns))
        ws2.cell(row=row_idx, column=6, value=info.sql_preview)
    ws2.freeze_panes = "A2"
    _autosize(ws2)

    ws3 = wb.create_sheet("SQL_Input")
    for col, h in enumerate (META_HEADERS, 1): 
        c = ws3.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT

    union_branches = {r.union_branch for r in records if r.union_branch is not None}
    max_depth = max((r. Lineage_depth for r in records), default=0)
    has_star = any(r.expression_type == "STAR" for r in records)
    output_col_count = len({r.output_col_position for r in records})

    meta_rows = [
        ("Dialect", dialect),
        ("Generated At", datetime.now().isoformat(timespec="seconds")),
        ("Total Output Columns", output_col_count),
        ("Total Source References (rows)", len(records)),
        ("CTE Count", len(cte_infos)),
        ("Max Lineage Depth", max_depth),
        ("Has UNIONS", "YES" if union_branches else "NO"),
        ("Union Branch Count", len (union_branches)),
        ("Has SELECT *", "YES" if has_star else "NO"),
        ("SQL", sql.strip()),
    ]
    for row_idx, (k, v) in enumerate(meta_rows, start=2):
        ws3.cell(row=row_idx, column=1, value=k)
        cell = ws3.cell(row=row_idx, column=2, value=v)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 100
    ws3.row_dimensions[len(meta_rows) + 1].height = 200

    wb. save(output_path)


#
# ORCHESTRATION
#

def _dedupe_records(records: list[LineageRecord]) -> list[LineageRecord]:
    """Drop functionally identical rows produced when the same column
    appears multiple times in one expression (e.g. `g - CASE WHEN r THEN g*0.1 ELSE g*0.05`
    refers to `g` 3 times, all resolving to the same source).
    """
    seen: set[tuple] = set()
    out: list[LineageRecord] = []
    for r in records:
        key = (
            r.output_col_position,
            r.output_col_name,
            r.expression_type,
            r.intermediate_cte_path,
            r.lineage_depth,
            r.union_branch,
            r.source_schema,
            r.source_table,
            r.source_column,
            r.source_table_alias,
            r.notes,
        )
        if key in seen: 
            continue
        seen.add(key)
        out.append(r)
    return out


def trace_lineage(
    sql: str,
    dialect: str = "default",
    output_path: str | None = None,
    schema_hint: dict[str, list[str]] | None = None,
) -> tuple[list[LineageRecord], list[CTEInfo]]:
    ast = parse_sql(sql, dialect)
    cte_map = extract_ctes(ast)
    top = get_top_level_select(ast)

    records: list[LineageRecord] = []
    if isinstance(top, exp.Union):
        for i, branch in enumerate(flatten_union(top), 1):
            records.extend(process_final_select(branch, cte_map, i, schema_hint))
    else:
        records.extend (process_final_select(top, cte_map, None, schema_hint))   
    records = _dedupe_records(records)
    cte_infos = analyze_cte_dependencies(cte_map)
    
    if output_path is not None:
        write_excel(records, cte_infos, output_path, sql, dialect)
    
    return records, cte_infos


#
# QUERY LOADER + MAIN
#

_Q_NAME_RE = re.compile(r"^q\d+$")


def load_queries(module_name: str) -> dict[str, str]: 
    mod = importlib.import_module(module_name)
    queries: dict[str, str] = {}
    for name in dir(mod):
        if not _Q_NAME_RE.match(name):
            continue
        val = getattr(mod, name)
        if isinstance(val, str) and val.strip():
            queries[name] = val
    return dict(sorted(queries.items(), key=lambda kv: int(kv[0][1:])))


def main() -> None:
    os.makedirs (OUTPUT_DIR, exist_ok=True)
    try:
        queries = load_queries(QUERIES_MODULE)
    except ModuleNotFoundError:
        print(f"[ERROR] Could not import module '{QUERIES_MODULE}'."
            f"Create {QUERIES_MODULE}.py with q1, q2, ... string variables.")
        return
    
    if not queries:
        print(f"[WARN] No queries found in {QUERIES_MODULE}.py " 
            f"(expected variables named q1, q2,..).")
        return

    successes = 0
    failures = 0
    for name, sql in queries.items():
        out_path = os.path.join(OUTPUT_DIR, f"{name}_lineage.xlsx")
        try:
            records, ctes = trace_lineage(sql, DIALECT, out_path, SCHEMA_HINT)
            print(f"[OK] {name} -> {out_path} "
                    f"({len(records)} rows, {len(ctes)} CTEs)")
            successes += 1
        except Exception as e:
            print(f"[FAIL] {name}: {type(e).__name__}: {e}")
            failures += 1
    
    print(f"\nDone. {successes} succeeded, {failures} failed. "
            f"Output dir: {OUTPUT_DIR}/")


if __name__  == "__main__":
    main()

